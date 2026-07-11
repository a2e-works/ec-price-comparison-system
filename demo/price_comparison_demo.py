"""
EC Price Comparison System (Portfolio Demo)

This is a simplified demonstration of the original workflow.
The production version includes API integration, error handling,
and business-specific logic which are not included in this repository.
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

# ----------------------------------
# Sample Price Lookup
# ----------------------------------

def get_amazon_price(jan):
    sample = {
        "490000000001": 3980,
        "490000000002": 2980,
        "490000000003": 4580,
    }
    return sample.get(jan)


def get_rakuten_price(jan):
    sample = {
        "490000000001": 3180,
        "490000000002": 2680,
        "490000000003": 3980,
    }
    return sample.get(jan)


def get_yahoo_price(jan):
    sample = {
        "490000000001": 3280,
        "490000000002": 2590,
        "490000000003": 4050,
    }
    return sample.get(jan)


# ----------------------------------
# Profit Calculation
# ----------------------------------

def calculate_profit_rate(sell_price, buy_price):

    if buy_price == 0:
        return 0

    return round(((sell_price - buy_price) / buy_price) * 100, 1)


# ----------------------------------
# Main Process
# ----------------------------------
BASE_DIR = Path(__file__).resolve().parent.parent

input_file = BASE_DIR / "sample" / "sample_input.csv"

df = pd.read_csv(input_file, dtype=str, encoding="cp932")

results = []

for _, row in df.iterrows():

    jan = row["JAN"]

    amazon = get_amazon_price(jan)
    rakuten = get_rakuten_price(jan)
    yahoo = get_yahoo_price(jan)

    if amazon is None:
        continue

    cheapest = min(rakuten, yahoo)

    profit = calculate_profit_rate(
        amazon,
        cheapest
    )

    results.append({

        "JAN": jan,
        "Amazon": amazon,
        "Rakuten": rakuten,
        "Yahoo": yahoo,
        "Purchase Price": cheapest,
        "Profit Rate (%)": profit

    })

output_file = BASE_DIR / "sample" / "sample_output.xlsx"

output = pd.DataFrame(results)

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    output.to_excel(
        writer,
        index=False,
        sheet_name="Profit Analysis"
    )

# Excel formatting
wb = load_workbook(output_file)
ws = wb["Profit Analysis"]

column_widths = {
    "A": 18,
    "B": 12,
    "C": 12,
    "D": 12,
    "E": 18,
    "F": 18,
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Freeze header row
ws.freeze_panes = "A2"

ws.auto_filter.ref = ws.dimensions

# Number formatting
for cell in ws["B"]:
    cell.number_format = '#,##0'

for cell in ws["C"]:
    cell.number_format = '#,##0'

for cell in ws["D"]:
    cell.number_format = '#,##0'

for cell in ws["E"]:
    cell.number_format = '#,##0'

for cell in ws["F"]:
    cell.number_format = '0.0'
wb.save(output_file)

print("Done.")
